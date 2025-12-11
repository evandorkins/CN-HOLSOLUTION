#!/usr/bin/env python3
"""
Holiday Credit Solution Documentation Generator
Parses UKG ProWFM JSON exports and creates Excel documentation with anomaly analysis
"""

import json
import os
from pathlib import Path
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

BASE_PATH = Path("fullHOLSolution")
# All objects now loaded from main folder
UPDATES_PATH = None
UPDATES_OBJECTS = []

# Header styling
HEADER_FONT = Font(bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
ANOMALY_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
WARNING_FILL = PatternFill(start_color="FFC000", end_color="FFC000", fill_type="solid")

def load_json(folder_name):
    """Load response.json from a folder - use updates folder for specific objects"""
    # Check if this object should come from updates
    if folder_name in UPDATES_OBJECTS:
        path = UPDATES_PATH / folder_name / "response.json"
        if path.exists():
            with open(path, 'r') as f:
                return json.load(f)
    # Otherwise use main folder
    path = BASE_PATH / folder_name / "response.json"
    if path.exists():
        with open(path, 'r') as f:
            return json.load(f)
    return None

def extract_items(data, object_key):
    """Extract items from the response structure"""
    items = []
    if not data:
        return items
    for response in data.get('itemsRetrieveResponses', []):
        title = response.get('itemDataInfo', {}).get('title', '')
        obj = response.get('responseObjectNode', {}).get(object_key, {})
        if obj:
            items.append({'title': title, 'data': obj})
    return items

def style_header(ws, row=1):
    """Style header row"""
    for cell in ws[row]:
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', wrap_text=True)

def auto_width(ws):
    """Auto-adjust column widths"""
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column].width = adjusted_width

# ================= EXTRACTION FUNCTIONS =================

def extract_accrual_codes():
    """Extract Accrual Codes (Banks)"""
    data = load_json("WSACfgAccrualCode")
    items = extract_items(data, "WSACfgAccrualCode")
    rows = []
    for item in items:
        d = item['data']
        rows.append({
            'Name': d.get('@Name', ''),
            'Abbreviation': d.get('@Abbreviation', ''),
            'Type': d.get('@Type', ''),
            'Hours Per Day': d.get('@HoursPerDay', ''),
            'Manual Edit': d.get('@ManualEdit', ''),
            'Display': d.get('@Display', ''),
            'Tracking Period': d.get('@TrackingPeriodName', '')
        })
    return pd.DataFrame(rows)

def extract_holidays():
    """Extract Holidays"""
    data = load_json("WSAHoliday")
    items = extract_items(data, "WSAHoliday")
    rows = []
    for item in items:
        d = item['data']
        name = d.get('@Name', '')
        display_name = d.get('@DisplayName', '')
        dates = d.get('WSAHolidayDates', {}).get('WSAHolidayDate', [])
        if isinstance(dates, dict):
            dates = [dates]
        date_count = len(dates)
        # Get first and last date
        if dates:
            first_date = dates[0].get('@StartDate', '')
            last_date = dates[-1].get('@StartDate', '')
        else:
            first_date = last_date = ''
        rows.append({
            'Name': name,
            'Display Name': display_name,
            'Date Count': date_count,
            'First Date': first_date,
            'Last Date': last_date
        })
    return pd.DataFrame(rows)

def extract_holiday_profiles():
    """Extract Holiday Profiles"""
    data = load_json("APIHolidayProfile")
    items = extract_items(data, "APIHolidayProfile")
    rows = []
    for item in items:
        d = item['data']
        name = d.get('@Name', '')
        active = d.get('@Active', '')
        dataset = d.get('HolidayProfileDataSet', {}).get('APIHolidayProfileDataSet', {})
        data_members = dataset.get('DataMembers', {}).get('APIHolidayProfileData', {})
        default_zone = data_members.get('@DefaultZoneRuleName', '')
        default_credit = data_members.get('@DefaultCreditRuleName', '')
        entries = data_members.get('Entries', {}).get('APIHolidayProfileDataEntry', [])
        if isinstance(entries, dict):
            entries = [entries]
        holiday_count = len(entries)

        rows.append({
            'Profile Name': name,
            'Active': active,
            'Default Zone Rule': default_zone,
            'Default Credit Rule': default_credit,
            'Holiday Count': holiday_count
        })
    return pd.DataFrame(rows)

def extract_holiday_profile_details():
    """Extract Holiday Profile Details (each holiday entry)"""
    data = load_json("APIHolidayProfile")
    items = extract_items(data, "APIHolidayProfile")
    rows = []
    for item in items:
        d = item['data']
        profile_name = d.get('@Name', '')
        dataset = d.get('HolidayProfileDataSet', {}).get('APIHolidayProfileDataSet', {})
        data_members = dataset.get('DataMembers', {}).get('APIHolidayProfileData', {})
        entries = data_members.get('Entries', {}).get('APIHolidayProfileDataEntry', [])
        if isinstance(entries, dict):
            entries = [entries]

        for entry in entries:
            rows.append({
                'Profile Name': profile_name,
                'Holiday Name': entry.get('@HolidayName', ''),
                'Zone Rule': entry.get('@ZoneRuleName', ''),
                'Credit Rule': entry.get('@CreditRuleName', ''),
                'Use Default Credit': entry.get('@UseDefaultCreditRule', ''),
                'Use Default Zone': entry.get('@UseDefaultZoneRule', '')
            })
    return pd.DataFrame(rows)

def extract_holiday_credit_rules():
    """Extract Holiday Credit Rules"""
    data = load_json("WSAHolidayCreditRule")
    items = extract_items(data, "WSAHolidayCreditRule")
    rows = []
    for item in items:
        d = item['data']
        name = d.get('@Name', '')

        # Eligibility settings
        elig = d.get('WSAEligibilityDeterminer', {}).get('WSAEligibilityDeterminer', {})
        before_holiday = elig.get('@BeforeHolidaySwitch', '')
        after_holiday = elig.get('@AfterHolidaySwitch', '')
        on_holiday = elig.get('@OnHolidaySwitch', '')
        either_before_after = elig.get('@EitherBeforeOrAfterHolidaySwitch', '')
        must_satisfy_all = elig.get('@MustSatisfyAllSwitch', '')
        scheduled_shift = elig.get('@ScheduledShiftTypeSwitch', '')
        alternate_rule = elig.get('@AlternateHolidayCreditRuleName', '')

        # Credit calculator settings
        calc = d.get('WSACreditCalculator', {}).get('WSACreditCalculator', {})
        contributing_shift = calc.get('@ContributingShiftName', '')
        credit_type = calc.get('@Type', '')
        fixed_amount = calc.get('@CreditFixedAmount', '')
        max_amount = calc.get('@MaxAmount', '')
        wage_calc = calc.get('@WageCalculation', '')
        as_if_worked = calc.get('@AsIfWorkedSwitch', '')

        # Pay codes
        credit_pay_codes = calc.get('CreditPayCodeNames', {})
        if isinstance(credit_pay_codes, dict):
            pcs = credit_pay_codes.get('SimpleValue', {})
            if isinstance(pcs, dict):
                credit_pay_code = pcs.get('@Value', '')
            elif isinstance(pcs, list):
                credit_pay_code = ', '.join([p.get('@Value', '') for p in pcs])
            else:
                credit_pay_code = ''
        else:
            credit_pay_code = ''

        ot_limit_pay_code = d.get('@OvertimeLimitPayCodeName', '')
        day_ot_limit_pay_code = d.get('@DayOvertimeLimitPayCodeName', '')
        ot_limit_type = d.get('@OvertimeLimitType', '')

        rows.append({
            'Name': name,
            'Before Holiday': before_holiday,
            'After Holiday': after_holiday,
            'On Holiday': on_holiday,
            'Either Before/After': either_before_after,
            'Must Satisfy All': must_satisfy_all,
            'Scheduled Shift Check': scheduled_shift,
            'Alternate Rule': alternate_rule,
            'Contributing Shift': contributing_shift,
            'Credit Type': credit_type,
            'Fixed Amount': fixed_amount,
            'Max Amount': max_amount,
            'Wage Calculation': wage_calc,
            'As If Worked': as_if_worked,
            'Credit Pay Code': credit_pay_code,
            'OT Limit Pay Code': ot_limit_pay_code,
            'Day OT Limit Pay Code': day_ot_limit_pay_code,
            'OT Limit Type': ot_limit_type
        })
    return pd.DataFrame(rows)

def extract_pay_codes():
    """Extract Pay Codes"""
    data = load_json("WSAPayCode")
    items = extract_items(data, "WSAPayCode")
    rows = []
    for item in items:
        d = item['data']
        rows.append({
            'Name': d.get('@Name', ''),
            'Type': d.get('@Type', ''),
            'Amount Type': d.get('@AmountType', ''),
            'Visible To User': d.get('@VisibleToUser', ''),
            'Visible In Report': d.get('@VisibleInReport', ''),
            'Visible In Main Area': d.get('@VisibleInMainArea', ''),
            'Timekeeping Only': d.get('@TimekeepingOnly', ''),
            'Payroll Only': d.get('@PayrollOnly', ''),
            'Edit Cnt To OT': d.get('@EditCntToOt', ''),
            'Edit Affect Shift Total': d.get('@EditAffShfTotal', ''),
            'Edit Excuse Absence': d.get('@EditExcuseAbsn', ''),
            'Schedule Hours Type': d.get('@ScheduleHoursType', ''),
            'Wage Multiply': d.get('@WageMultiply', ''),
            'Wage Addition': d.get('@WageAddition', '')
        })
    return pd.DataFrame(rows)

def extract_contributing_pay_code_rules():
    """Extract Contributing Pay Code Rules"""
    data = load_json("WSAContributingPayCodeRule")
    items = extract_items(data, "WSAContributingPayCodeRule")
    rows = []
    for item in items:
        d = item['data']
        pay_codes = d.get('PayCodeNames', {}).get('SimpleValue', {})
        if isinstance(pay_codes, dict):
            pc_value = pay_codes.get('@Value', '')
        elif isinstance(pay_codes, list):
            pc_value = ', '.join([p.get('@Value', '') for p in pay_codes])
        else:
            pc_value = ''
        rows.append({
            'Name': d.get('@Name', ''),
            'Is Prep Payroll': d.get('@IsPrepPayrollSw', ''),
            'Is Contribute Shift': d.get('@IsContributeShftSw', ''),
            'Pay Codes': pc_value
        })
    return pd.DataFrame(rows)

def extract_contributing_shift_rules():
    """Extract Contributing Shift Rules"""
    data = load_json("WSAContributingShiftRule")
    items = extract_items(data, "WSAContributingShiftRule")
    rows = []
    for item in items:
        d = item['data']
        rows.append({
            'Name': d.get('@Name', ''),
            'Contributing Pay Code Rule': d.get('@ContributingPayCodeRuleName', ''),
            'Look Back Type': d.get('@LookBackType', ''),
            'Look Back Time': d.get('@LookBackTime', ''),
            'Look Back Unit': d.get('@LookBackUnit', ''),
            'Days Of Week Include Type': d.get('@DaysOfWeekIncludeType', ''),
            'Minimum Shift Length': d.get('@MinimumShiftLength', ''),
            'Include Complete Period': d.get('@IncludeCompletePeriod', '')
        })
    return pd.DataFrame(rows)

def extract_balance_cascades():
    """Extract Balance Cascades"""
    data = load_json("WSABalanceCascade")
    items = extract_items(data, "WSABalanceCascade")
    rows = []
    for item in items:
        d = item['data']
        cascade_items = d.get('BalanceCascadeItems', {}).get('WSABalanceCascadeItem', {})
        if isinstance(cascade_items, dict):
            payout_code = cascade_items.get('@PayoutToPayCode', '')
            deduct_from = cascade_items.get('@DeductFrom', '')
            transfer_to = cascade_items.get('@TransferToAccrualCode', '')
        else:
            payout_code = deduct_from = transfer_to = ''

        rows.append({
            'Name': d.get('@Name', ''),
            'Accrual Code': d.get('@AccrualCode', ''),
            'Description': d.get('@Description', ''),
            'Date Pattern': d.get('@DatePattern', ''),
            'Type': d.get('@Type', ''),
            'Fixed Amount': d.get('@FixedAmount', ''),
            'Reduce To Zero': d.get('@ReduceToZero', ''),
            'Payout To Pay Code': payout_code,
            'Deduct From': deduct_from,
            'Transfer To Accrual': transfer_to
        })
    return pd.DataFrame(rows)

def extract_balance_cascade_groups():
    """Extract Balance Cascade Groups"""
    data = load_json("WSABalanceCascadeGroup")
    items = extract_items(data, "WSABalanceCascadeGroup")
    rows = []
    for item in items:
        d = item['data']
        cascades = d.get('SelectedBalanceCascades', {}).get('WSASelectedBalanceCascades', [])
        if isinstance(cascades, dict):
            cascades = [cascades]
        cascade_names = ', '.join([c.get('@Name', '') for c in cascades])

        rows.append({
            'Name': d.get('@Name', ''),
            'Description': d.get('@Description', ''),
            'Cascade Count': len(cascades),
            'Cascades': cascade_names
        })
    return pd.DataFrame(rows)

def extract_limits():
    """Extract Limits"""
    data = load_json("WSALimit")
    items = extract_items(data, "WSALimit")
    rows = []
    for item in items:
        d = item['data']
        rows.append({
            'Name': d.get('@Name', ''),
            'Limit Type': d.get('@LimitType', ''),
            'Date Pattern': d.get('@DatePattern', ''),
            'Max Limit': d.get('@MaxLimit', ''),
            'Forgiven Balance': d.get('@ForgivenBalance', '')
        })
    return pd.DataFrame(rows)

def extract_date_patterns():
    """Extract Date Patterns"""
    data = load_json("WSADatePattern")
    items = extract_items(data, "WSADatePattern")
    rows = []
    for item in items:
        d = item['data']
        finder = d.get('DateFinder', {}).get('WSADateFinder', {})
        interval = d.get('Interval', {}).get('WSAInterval', {})

        rows.append({
            'Name': d.get('@Name', ''),
            'Expected Hours': d.get('@ExpectedHours', ''),
            'Offset Amount': d.get('@OffsetAmount', ''),
            'Offset Date Name': d.get('@OffsetDateName', ''),
            'Custom Date Type': finder.get('@CustomDateType', ''),
            'Date Finder Type': finder.get('@DateFinderType', ''),
            'Time Unit Type': finder.get('@TimeUnitType', ''),
            'Num Intervals': interval.get('@NumIntervals', ''),
            'Interval Time Unit': interval.get('@TimeUnitType', '')
        })
    return pd.DataFrame(rows)

def extract_custom_dates():
    """Extract Custom Dates"""
    data = load_json("WSACustomDate")
    items = extract_items(data, "WSACustomDate")
    rows = []
    for item in items:
        d = item['data']
        rows.append({
            'Name': d.get('@Name', ''),
            'Date': d.get('@Date', ''),
            'Site Wide': d.get('@SiteWide', ''),
            'Reference Custom Date': d.get('@ReferenceCustomDate', '')
        })
    return pd.DataFrame(rows)

def extract_employment_terms():
    """Extract Employment Terms"""
    data = load_json("EmploymentTerm")
    items = extract_items(data, "EmploymentTerm")
    rows = []
    for item in items:
        d = item['data']
        versions = d.get('Versions', {}).get('EmploymentTermVersion', {})
        if isinstance(versions, list):
            versions = versions[0] if versions else {}

        rows.append({
            'Name': d.get('@Name', ''),
            'Description': d.get('@Description', ''),
            'Is Active': d.get('@IsActive', ''),
            'Allows Inheritance': d.get('@AllowsInheritance', ''),
            'Holiday Profile': versions.get('@HolidayProfile', ''),
            'Cascade Profile': versions.get('@CascadeProfile', ''),
            'Accrual Profile': versions.get('@AccrualProfile', ''),
            'Time Off Rule': versions.get('@TimeOffRule', ''),
            'Pay Rule': versions.get('@PayRule', '')
        })
    return pd.DataFrame(rows)

def extract_accrual_policies():
    """Extract Accrual Policies"""
    data = load_json("WSAAccrualPolicy")
    items = extract_items(data, "WSAAccrualPolicy")
    rows = []
    for item in items:
        d = item['data']
        name = d.get('@Name', '')
        eff = d.get('EffectiveAccrualPolicies', {}).get('WSAEffectiveAccrualPolicy', {})
        if isinstance(eff, list):
            eff = eff[0] if eff else {}
        if isinstance(eff, str):
            eff = {}

        accrual_code = eff.get('@AccrualCode', '') if isinstance(eff, dict) else ''
        payout_code = eff.get('@AccrualPayoutPayCode', '') if isinstance(eff, dict) else ''
        overdraft = eff.get('@OverdraftError', '') if isinstance(eff, dict) else ''

        # Get limits
        limits_container = eff.get('Limits', {}) if isinstance(eff, dict) else {}
        if isinstance(limits_container, dict):
            limits = limits_container.get('WSAAccrualPolicyLimit', [])
        else:
            limits = []
        if isinstance(limits, dict):
            limits = [limits]
        limit_rules = ', '.join([l.get('@GrantRule', '') for l in limits if isinstance(l, dict)])

        # Get taking limits
        taking_container = eff.get('TakingLimits', {}) if isinstance(eff, dict) else {}
        if isinstance(taking_container, dict):
            taking = taking_container.get('WSAAccrualPolicyTakingLimit', {})
        else:
            taking = {}
        taking_limit = taking.get('@DisallowAmount', '') if isinstance(taking, dict) else ''

        rows.append({
            'Name': name,
            'Accrual Code': accrual_code,
            'Payout Pay Code': payout_code,
            'Overdraft Error': overdraft,
            'Taking Limit': taking_limit,
            'Limit Rules': limit_rules
        })
    return pd.DataFrame(rows)

def extract_accrual_profiles():
    """Extract Accrual Profiles"""
    data = load_json("WSAAccrualProfile")
    items = extract_items(data, "WSAAccrualProfile")
    rows = []
    for item in items:
        d = item['data']
        name = d.get('@Name', '')

        # Get policies
        policies = d.get('AccrualPolicies', {}).get('AccrualPolicyName', [])
        if isinstance(policies, str):
            policies = [policies]
        elif isinstance(policies, dict):
            policies = [policies.get('@Name', '')]
        elif isinstance(policies, list):
            policies = [p if isinstance(p, str) else p.get('@Name', '') for p in policies]

        # Get cascade group
        cascade_group = d.get('@BalanceCascadeGroupName', '')

        rows.append({
            'Name': name,
            'Policy Count': len(policies),
            'Cascade Group': cascade_group,
            'Policies': ', '.join(policies[:5]) + ('...' if len(policies) > 5 else '')
        })
    return pd.DataFrame(rows)

# ================= ANOMALY DETECTION =================

def detect_anomalies():
    """Detect configuration anomalies and inconsistencies"""
    findings = []

    # 1. Check Pay Codes for inconsistencies
    pc_df = extract_pay_codes()
    if not pc_df.empty:
        # Check for naming convention issues
        holiday_codes = pc_df[pc_df['Name'].str.contains('HOL|HOLIDAY|H-', case=False, na=False)]

        # Check for zz prefix (test codes)
        test_codes = pc_df[pc_df['Name'].str.startswith('zz', na=False)]
        if not test_codes.empty:
            findings.append({
                'Category': 'Pay Codes',
                'Severity': 'Warning',
                'Finding': f"Found {len(test_codes)} test pay codes with 'zz' prefix",
                'Details': ', '.join(test_codes['Name'].tolist()),
                'Recommendation': 'Review if test codes should be removed from production'
            })

        # Check for inconsistent Type settings
        credit_codes = pc_df[pc_df['Name'].str.contains('CREDIT', case=False, na=False)]
        if not credit_codes.empty:
            type_counts = credit_codes['Type'].value_counts()
            if len(type_counts) > 1:
                findings.append({
                    'Category': 'Pay Codes',
                    'Severity': 'Medium',
                    'Finding': 'Inconsistent Type settings among CREDIT pay codes',
                    'Details': f"Types found: {dict(type_counts)}",
                    'Recommendation': 'Verify all CREDIT codes should have same Type'
                })

        # Check Timekeeping Only inconsistencies
        holiday_credit_codes = pc_df[pc_df['Name'].str.contains('HOLIDAY CREDIT', case=False, na=False)]
        if not holiday_credit_codes.empty:
            tk_counts = holiday_credit_codes['Timekeeping Only'].value_counts()
            if len(tk_counts) > 1:
                findings.append({
                    'Category': 'Pay Codes',
                    'Severity': 'Medium',
                    'Finding': 'Inconsistent Timekeeping Only setting among HOLIDAY CREDIT codes',
                    'Details': f"Settings: {dict(tk_counts)}",
                    'Recommendation': 'Verify all HOLIDAY CREDIT codes have consistent TK settings'
                })

    # 2. Check Holiday Credit Rules
    hcr_df = extract_holiday_credit_rules()
    if not hcr_df.empty:
        # Check for test/forfeited rules
        test_rules = hcr_df[hcr_df['Name'].str.contains('zz|TEST', case=False, na=False)]
        if not test_rules.empty:
            findings.append({
                'Category': 'Holiday Credit Rules',
                'Severity': 'Warning',
                'Finding': f"Found {len(test_rules)} test/zz prefixed credit rules",
                'Details': ', '.join(test_rules['Name'].tolist()),
                'Recommendation': 'Review if test rules should be removed'
            })

        # Check eligibility consistency - Exempt vs Non-Exempt
        exempt_rules = hcr_df[hcr_df['Name'].str.startswith('Exempt ', na=False)]
        non_exempt_rules = hcr_df[hcr_df['Name'].str.startswith('Non Exempt ', na=False)]

        # Exempt should NOT require before/after holiday
        if not exempt_rules.empty:
            exempt_with_eligibility = exempt_rules[
                (exempt_rules['Before Holiday'] == True) | (exempt_rules['After Holiday'] == True)
            ]
            if not exempt_with_eligibility.empty:
                findings.append({
                    'Category': 'Holiday Credit Rules',
                    'Severity': 'Info',
                    'Finding': f"{len(exempt_with_eligibility)} Exempt rules have Before/After Holiday checks enabled",
                    'Details': ', '.join(exempt_with_eligibility['Name'].tolist()[:5]),
                    'Recommendation': 'Verify this is intentional for exempt employees'
                })

        # Non-Exempt should typically require before/after
        if not non_exempt_rules.empty:
            non_exempt_without = non_exempt_rules[
                (non_exempt_rules['Before Holiday'] == False) & (non_exempt_rules['After Holiday'] == False)
            ]
            if not non_exempt_without.empty and len(non_exempt_without) < len(non_exempt_rules):
                findings.append({
                    'Category': 'Holiday Credit Rules',
                    'Severity': 'Medium',
                    'Finding': 'Inconsistent eligibility settings among Non-Exempt rules',
                    'Details': f"{len(non_exempt_without)} of {len(non_exempt_rules)} don't check Before/After",
                    'Recommendation': 'Verify FORFEITED rules intentionally skip eligibility'
                })

        # Check ScheduledShiftTypeSwitch inconsistency
        labor_rules = hcr_df[hcr_df['Name'].str.contains('LABOR', case=False, na=False)]
        if not labor_rules.empty:
            sched_check = labor_rules['Scheduled Shift Check'].value_counts()
            if len(sched_check) > 1:
                findings.append({
                    'Category': 'Holiday Credit Rules',
                    'Severity': 'High',
                    'Finding': 'LABOR DAY rules have inconsistent Scheduled Shift Check settings',
                    'Details': f"Settings: {dict(sched_check)} - {labor_rules[['Name', 'Scheduled Shift Check']].to_dict('records')}",
                    'Recommendation': 'All LABOR DAY rules should have consistent scheduled shift settings'
                })

    # 3. Check Contributing Shift Rules
    csr_df = extract_contributing_shift_rules()
    if not csr_df.empty:
        # Check for wrong contributing pay code rule reference
        forfeit_presidents = csr_df[csr_df['Name'] == 'FORFEIT PRESIDENTS DAY']
        if not forfeit_presidents.empty:
            cp_rule = forfeit_presidents['Contributing Pay Code Rule'].values[0]
            if 'MLK' in cp_rule:
                findings.append({
                    'Category': 'Contributing Shift Rules',
                    'Severity': 'High',
                    'Finding': 'FORFEIT PRESIDENTS DAY uses wrong Contributing Pay Code Rule',
                    'Details': f"Currently references '{cp_rule}' - should reference PRESIDENTS",
                    'Recommendation': 'Update to use H-FORFEIT PENDING PRESIDENTS'
                })

    # 4. Check Balance Cascades
    bc_df = extract_balance_cascades()
    if not bc_df.empty:
        # Check for date pattern inconsistencies
        payout_30 = bc_df[bc_df['Name'].str.contains('Plus 30', case=False, na=False)]
        if not payout_30.empty:
            # Check Thanksgiving specifically
            thanksgiving_30 = payout_30[payout_30['Name'].str.contains('Thanksgiving', case=False, na=False)]
            if not thanksgiving_30.empty:
                date_pattern = thanksgiving_30['Date Pattern'].values[0]
                if 'Plus 30' not in date_pattern:
                    findings.append({
                        'Category': 'Balance Cascades',
                        'Severity': 'High',
                        'Finding': 'Thanksgiving Plus 30 cascade uses wrong date pattern',
                        'Details': f"Uses '{date_pattern}' instead of 'Plus 30' pattern",
                        'Recommendation': 'Update date pattern to HOL-Thanksgiving Day Plus 30 Annual'
                    })

    # 5. Check Date Patterns
    dp_df = extract_date_patterns()
    if not dp_df.empty:
        # Check for missing offset on Annual patterns
        annual_patterns = dp_df[dp_df['Name'].str.contains('Annual', na=False)]
        no_offset = annual_patterns[annual_patterns['Offset Amount'].isna() | (annual_patterns['Offset Amount'] == '')]
        if not no_offset.empty:
            findings.append({
                'Category': 'Date Patterns',
                'Severity': 'Low',
                'Finding': f"{len(no_offset)} Annual date patterns missing offset",
                'Details': ', '.join(no_offset['Name'].tolist()),
                'Recommendation': 'Verify if offset should be specified'
            })

        # Check for special Christmas pattern
        christmas_minus = dp_df[dp_df['Name'].str.contains('CHRISTMAS DAY MINUS 2|Christmas Day Minus 2', na=False, regex=True)]
        if christmas_minus.empty:
            # Check for zz pattern
            zz_christmas = dp_df[dp_df['Name'].str.contains('zz CHRISTMAS', case=False, na=False)]
            if not zz_christmas.empty:
                findings.append({
                    'Category': 'Date Patterns',
                    'Severity': 'Warning',
                    'Finding': 'Christmas cascade uses test/zz date pattern',
                    'Details': f"Found: {', '.join(zz_christmas['Name'].tolist())}",
                    'Recommendation': 'Review if this should use production date pattern'
                })

    # 6. Check Holiday Profiles
    hp_df = extract_holiday_profiles()
    hp_details = extract_holiday_profile_details()
    if not hp_details.empty:
        # Check for inconsistent credit rules within same profile
        for profile in hp_details['Profile Name'].unique():
            profile_entries = hp_details[hp_details['Profile Name'] == profile]
            credit_rules = profile_entries['Credit Rule'].unique()

            # Check Non Exempt PT profile specifically
            if 'Non Exempt PT' in profile:
                veteran_entry = profile_entries[profile_entries['Holiday Name'].str.contains("Veteran", case=False, na=False)]
                if not veteran_entry.empty:
                    credit_rule = veteran_entry['Credit Rule'].values[0]
                    if credit_rule == 'Non Exempt' and 'PT' not in credit_rule:
                        findings.append({
                            'Category': 'Holiday Profiles',
                            'Severity': 'High',
                            'Finding': f"Profile '{profile}' has inconsistent credit rule for Veterans Day",
                            'Details': f"Veterans Day uses '{credit_rule}' while others use PT variant",
                            'Recommendation': 'Update Veterans Day to use Non Exempt PT credit rule'
                        })

            # Check for profile using wrong FTE credit rule
            if 'PT 0.6' in profile:
                memorial_entry = profile_entries[profile_entries['Holiday Name'].str.contains("Memorial", case=False, na=False)]
                if not memorial_entry.empty:
                    credit_rule = memorial_entry['Credit Rule'].values[0]
                    if '0.4' in credit_rule or 'PT 40' in credit_rule:
                        findings.append({
                            'Category': 'Holiday Profiles',
                            'Severity': 'High',
                            'Finding': f"Profile '{profile}' has wrong FTE credit rule for Memorial Day",
                            'Details': f"Uses '{credit_rule}' instead of 0.6 variant",
                            'Recommendation': 'Update to correct FTE percentage rule'
                        })

    # 7. Check Accrual Codes consistency
    ac_df = extract_accrual_codes()
    if not ac_df.empty:
        # All HOL codes should have same settings
        hol_codes = ac_df[ac_df['Name'].str.startswith('HOL-', na=False)]
        if not hol_codes.empty:
            for col in ['Type', 'Hours Per Day', 'Manual Edit', 'Display']:
                unique_vals = hol_codes[col].unique()
                if len(unique_vals) > 1:
                    findings.append({
                        'Category': 'Accrual Codes',
                        'Severity': 'Medium',
                        'Finding': f"Inconsistent {col} setting among HOL accrual codes",
                        'Details': f"Values found: {list(unique_vals)}",
                        'Recommendation': 'Verify all HOL codes should have same settings'
                    })

    # 8. Check Holiday dates consistency
    hol_df = extract_holidays()
    if not hol_df.empty:
        # Check for holidays without enough future dates
        low_dates = hol_df[hol_df['Date Count'] < 3]
        if not low_dates.empty:
            findings.append({
                'Category': 'Holidays',
                'Severity': 'Warning',
                'Finding': f"{len(low_dates)} holidays have fewer than 3 dates defined",
                'Details': ', '.join(low_dates['Name'].tolist()),
                'Recommendation': 'Consider adding future year dates'
            })

        # Check naming convention (ADVS vs regular)
        advs_holidays = hol_df[hol_df['Name'].str.startswith('ADVS', na=False)]
        if not advs_holidays.empty:
            findings.append({
                'Category': 'Holidays',
                'Severity': 'Info',
                'Finding': f"Found {len(advs_holidays)} ADVS-prefixed holidays",
                'Details': 'These appear to be alternate/specific holiday definitions',
                'Recommendation': 'Verify ADVS holidays are correctly used in profiles'
            })

        # Check for missing standard holidays (non-ADVS)
        expected_holidays = [
            'Independence Day', 'Labor Day', 'Martin Luther King', 'MLK',
            'Memorial Day', 'Presidents Day', 'Veterans Day'
        ]
        non_advs = hol_df[~hol_df['Name'].str.startswith('ADVS', na=False)]
        non_advs_names = ' '.join(non_advs['Name'].tolist()).upper()
        missing = []
        for hol in expected_holidays:
            if hol.upper() not in non_advs_names:
                missing.append(hol)
        if missing:
            findings.append({
                'Category': 'Holidays',
                'Severity': 'High',
                'Finding': 'Holiday export appears incomplete - missing standard holidays',
                'Details': f"Only found {len(non_advs)} non-ADVS holidays. Missing: Independence Day, Labor Day, MLK Day, Memorial Day, Presidents Day, Veterans Day",
                'Recommendation': 'Re-export holidays from UKG ProWFM to include all standard holidays'
            })

    # 9. Check Holiday Credit Rules for test contributing shifts
    hcr_df = extract_holiday_credit_rules()
    if not hcr_df.empty:
        test_contrib = hcr_df[hcr_df['Contributing Shift'].str.contains('zz|TEST', case=False, na=False)]
        if not test_contrib.empty:
            findings.append({
                'Category': 'Holiday Credit Rules',
                'Severity': 'Warning',
                'Finding': f"{len(test_contrib)} Holiday Credit Rules use test Contributing Shift",
                'Details': f"Rules using 'zz TEST HOL FORFEIT PENDING': {', '.join(test_contrib['Name'].tolist()[:5])}...",
                'Recommendation': 'Verify FORFEITED rules should use actual FORFEIT contributing shift rules'
            })

    return pd.DataFrame(findings)

# ================= MAIN EXECUTION =================

def main():
    print("Generating Holiday Credit Solution Documentation...")

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    # Create sheets for each object type
    sheets_data = [
        ("Accrual Codes", extract_accrual_codes()),
        ("Holidays", extract_holidays()),
        ("Holiday Profiles", extract_holiday_profiles()),
        ("Holiday Profile Details", extract_holiday_profile_details()),
        ("Holiday Credit Rules", extract_holiday_credit_rules()),
        ("Pay Codes", extract_pay_codes()),
        ("Contrib Pay Code Rules", extract_contributing_pay_code_rules()),
        ("Contrib Shift Rules", extract_contributing_shift_rules()),
        ("Balance Cascades", extract_balance_cascades()),
        ("Balance Cascade Groups", extract_balance_cascade_groups()),
        ("Limits", extract_limits()),
        ("Date Patterns", extract_date_patterns()),
        ("Custom Dates", extract_custom_dates()),
        ("Employment Terms", extract_employment_terms()),
        ("Accrual Policies", extract_accrual_policies()),
        ("Accrual Profiles", extract_accrual_profiles()),
    ]

    # Build TOC data (sheet name and row count)
    toc_data = []
    for sheet_name, df in sheets_data:
        if not df.empty:
            toc_data.append((sheet_name, len(df)))

    # Summary sheet
    ws['A1'] = "Holiday Credit Solution Documentation"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A3'] = "Generated from UKG ProWFM JSON exports"
    ws['A4'] = f"Object Types Documented: {len(toc_data)}"

    # Table of Contents
    ws['A6'] = "Table of Contents"
    ws['A6'].font = Font(bold=True, size=12)
    ws['A7'] = "Sheet Name"
    ws['B7'] = "Object Count"
    ws['A7'].font = HEADER_FONT
    ws['B7'].font = HEADER_FONT
    ws['A7'].fill = HEADER_FILL
    ws['B7'].fill = HEADER_FILL

    for idx, (sheet_name, count) in enumerate(toc_data, start=8):
        ws[f'A{idx}'] = sheet_name
        ws[f'B{idx}'] = count

    # Add ANOMALIES row placeholder (will update after detecting)
    anomalies_row = 8 + len(toc_data)
    ws[f'A{anomalies_row}'] = "ANOMALIES & FINDINGS"
    ws[f'B{anomalies_row}'] = 0  # Will update later

    # Auto-width for summary columns
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 15

    for sheet_name, df in sheets_data:
        if df.empty:
            print(f"  Skipping {sheet_name} - no data")
            continue

        print(f"  Creating sheet: {sheet_name} ({len(df)} rows)")
        ws = wb.create_sheet(title=sheet_name[:31])  # Excel limit

        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=value)

        style_header(ws)
        auto_width(ws)

    # Add Anomalies sheet
    print("  Detecting anomalies...")
    anomalies_df = detect_anomalies()
    if not anomalies_df.empty:
        ws_anomalies = wb.create_sheet(title="ANOMALIES & FINDINGS")
        for r_idx, row in enumerate(dataframe_to_rows(anomalies_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                cell = ws_anomalies.cell(row=r_idx, column=c_idx, value=value)
                if r_idx > 1:
                    severity = anomalies_df.iloc[r_idx-2]['Severity'] if r_idx-2 < len(anomalies_df) else ''
                    if severity == 'High':
                        cell.fill = WARNING_FILL
                    elif severity in ['Medium', 'Warning']:
                        cell.fill = ANOMALY_FILL

        style_header(ws_anomalies)
        auto_width(ws_anomalies)
        print(f"  Found {len(anomalies_df)} anomalies/findings")

        # Update anomalies count in Summary TOC
        summary_ws = wb["Summary"]
        summary_ws[f'B{anomalies_row}'] = len(anomalies_df)

    # Save workbook
    output_file = "HolidayCreditSolution_Documentation.xlsx"
    wb.save(output_file)
    print(f"\nDocumentation saved to: {output_file}")

    # Print anomalies summary
    if not anomalies_df.empty:
        print("\n" + "="*60)
        print("CONFIGURATION ANOMALIES & FINDINGS SUMMARY")
        print("="*60)
        for _, row in anomalies_df.iterrows():
            severity_marker = "🔴" if row['Severity'] == 'High' else "🟡" if row['Severity'] in ['Medium', 'Warning'] else "🔵"
            print(f"\n{severity_marker} [{row['Severity']}] {row['Category']}")
            print(f"   Finding: {row['Finding']}")
            print(f"   Details: {row['Details'][:100]}...")
            print(f"   Recommendation: {row['Recommendation']}")

if __name__ == "__main__":
    main()
